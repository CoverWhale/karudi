import os
import pandas as pd
import polars as pl
import clickhouse_connect
from datetime import datetime
import argparse
from functools import reduce
import operator
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from dq_table_map import build_table_pairs

# ------------------------- ARGUMENTS ------------------------- #

def parse_args():
    parser = argparse.ArgumentParser(description="Run DQ checks for a given model and optional date.")
    parser.add_argument("--model", required=True, choices=["WIP", "Quotes", "Policy"], help="Data model to check")
    parser.add_argument("--date", help="Date folder to run against (format: YYYY-MM-DD)")
    return parser.parse_args()

# ------------------------- CONFIG ------------------------- #

INCLUDE_ROW_DIFFS = True
TIMESTAMPED_REPORT = True
OUTPUT_DIR = "mismatch_reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

client = clickhouse_connect.get_client(host="localhost")

# ------------------------- HELPERS ------------------------- #

def clean_cw_string_columns(df: pl.DataFrame) -> pl.DataFrame:
    def strip_trailing_zero_str(x):
        try:
            float_val = float(x)
            if float_val.is_integer():
                return str(int(float_val))
            else:
                return str(float_val)
        except:
            return x
    return df.with_columns([
        df[col].map_elements(strip_trailing_zero_str).alias(col)
        for col in df.columns
        if df[col].dtype == pl.Utf8
    ])

def fetch_table(table_name, filter_clause=None):
    query = f"SELECT * FROM {table_name}"
    if filter_clause:
        query += f" WHERE {filter_clause}"
    result = client.query(query)
    return pl.DataFrame(result.result_rows, schema=result.column_names)

# ------------------------- CORE COMPARISON ------------------------- #

def compare_tables(name, df_a, df_b, join_keys, financial_fields):
    for a_key, b_key in join_keys.items():
        if a_key != b_key:
            df_b = df_b.rename({b_key: a_key})

    df_merged = df_a.join(df_b, on=list(join_keys.keys()), how="left", suffix="_b")

    # Normalize mismatched types before comparison
    for col in df_a.columns:
        if col in df_b.columns and col not in join_keys.keys():
            a_dtype = df_a[col].dtype
            b_dtype = df_b[col].dtype
            if a_dtype != b_dtype:
                df_merged = df_merged.with_columns([
                    df_merged[col].cast(pl.Utf8),
                    df_merged[f"{col}_b"].cast(pl.Utf8)
                ])

    non_key_b_cols = [c for c in df_b.columns if c not in join_keys.values()]
    match_indicator_col = f"{non_key_b_cols[0]}_b" if non_key_b_cols else f"{list(join_keys.keys())[0]}_b"
    matched = df_merged.filter(~pl.col(match_indicator_col).is_null()).shape[0]
    total_a = df_a.shape[0]
    match_pct = (matched / total_a) * 100 if total_a else 0
    print(f"[{name}] ✅ {matched}/{total_a} ({match_pct:.2f}%) records matched.")

    mismatch_summary = []
    column_stats_rows = []
    false_mismatch_rows = []
    matched_col_count = 0
    common_cols = [c for c in df_a.columns if c in df_b.columns and c not in join_keys.keys()]

    for col in common_cols:
        col_a_raw = df_merged[col]
        col_b_raw = df_merged[f"{col}_b"]

        col_a = pl.col(col)
        col_b = pl.col(f"{col}_b")

        # True mismatch: exclude cases where CW value is null or empty
        mismatch_filter = ((col_a != col_b) & ~col_a.is_null() & (col_a != "") & ~col_b.is_null())
        mismatch_count = df_merged.filter(mismatch_filter).shape[0]

        # False mismatches (CW is null or empty, Solartis is not)
        false_mismatch_filter = ((col_a.is_null() | (col_a == "")) & ~col_b.is_null() & (col_b != ""))
        false_mismatch_count = df_merged.filter(false_mismatch_filter).shape[0]

        total = df_merged.shape[0]
        match_percent = ((total - mismatch_count - false_mismatch_count) / total) * 100 if total else 0
        column_stats_rows.append({
            "column_name": col,
            "mismatches": mismatch_count,
            "false_mismatch_count": false_mismatch_count,
            "match_percent": round(match_percent, 2)
        })
        mismatch_summary.append((col, mismatch_count))
        if mismatch_count == 0:
            matched_col_count += 1

        if false_mismatch_count > 0:
            false_mismatch_rows.append({
                "table_name": name,
                "column_name": col,
                "false_mismatch_count": false_mismatch_count
            })

    summary_df = pl.DataFrame(mismatch_summary, schema=["column_name", "mismatches"]).sort("mismatches", descending=True)
    column_stats_df = pd.DataFrame(column_stats_rows)
    false_mismatch_df = pd.DataFrame(false_mismatch_rows)

    # Vertical row diffs
    row_diff_df = pl.DataFrame()
    if INCLUDE_ROW_DIFFS and common_cols:
        mismatch_filter = reduce(operator.or_, [
            ((pl.col(c) != pl.col(f"{c}_b")) & ~pl.col(c).is_null() & (pl.col(c) != "") & ~pl.col(f"{c}_b").is_null())
            for c in common_cols
        ])
        mismatches = df_merged.filter(mismatch_filter).limit(100)
        vertical_rows = []
        for row in mismatches.iter_rows(named=True):
            key_values = {key: row[key] for key in join_keys.keys()}
            for col in common_cols:
                a_val = row.get(col)
                b_val = row.get(f"{col}_b")
                vertical_rows.append({
                    **{k: str(v) for k, v in key_values.items()},
                    "ColumnName": col,
                    "CW_Value": str(a_val),
                    "Solartis_Value": str(b_val)
                })
        row_diff_df = pl.DataFrame(vertical_rows)

    # Financials
    financial_rows = []
    for field in financial_fields:
        if field in df_merged.columns and f"{field}_b" in df_merged.columns:
            a_sum = df_merged[field].fill_null(0).sum()
            b_sum = df_merged[f"{field}_b"].fill_null(0).sum()
            diff = a_sum - b_sum
            pct_diff = (diff / b_sum * 100) if b_sum != 0 else None
            financial_rows.append({
                "field": field,
                "table_a_sum": round(a_sum, 2),
                "table_b_sum": round(b_sum, 2),
                "difference": round(diff, 2),
                "percent_difference": round(pct_diff, 2) if pct_diff is not None else "N/A"
            })
    financial_df = pd.DataFrame(financial_rows)

    return {
        "summary_df": summary_df,
        "row_diff_df": row_diff_df,
        "financial_df": financial_df,
        "column_stats_df": column_stats_df,
        "false_mismatch_df": false_mismatch_df,
        "matched": matched,
        "total_a": total_a,
        "match_pct": match_pct,
        "matched_columns": matched_col_count,
        "total_columns": len(common_cols)
    }

# ------------------------- FORMATTING ------------------------- #

def apply_excel_formatting(filepath):
    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith("_financials"):
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if abs(cell.value) > 5:
                            cell.fill = PatternFill(start_color="FF9999", fill_type="solid")
                        elif abs(cell.value) > 1:
                            cell.fill = PatternFill(start_color="FFF799", fill_type="solid")
    wb.save(filepath)

# ------------------------- MAIN ------------------------- #

def main():
    args = parse_args()
    model = args.model
    date_folder = args.date or datetime.now().strftime("%Y-%m-%d")
    print(f"\n🚀 Starting DQ checks for model: `{model}`, date folder: `{date_folder}`")

    table_pairs = build_table_pairs(model)
    all_sheets = {}
    summary_records = []
    combined_false_mismatches = pd.DataFrame()

    for pair in table_pairs:
        try:
            name = pair["name"]
            print(f"\n🔍 Comparing: {name}")
            table_a_filter = pair.get("filters", {}).get("table_a")
            table_b_filter = pair.get("filters", {}).get("table_b")
            df_a_raw = fetch_table(pair["table_a"], filter_clause=table_a_filter)
            df_a = clean_cw_string_columns(df_a_raw)
            df_b = fetch_table(pair["table_b"], filter_clause=table_b_filter)

            result = compare_tables(name, df_a, df_b, pair["join_keys"], pair.get("financial_fields", []))

            summary_records.append({
                "table_name": name,
                "rows_in_cw_table": df_a.shape[0],
                "rows_in_solartis_table": df_b.shape[0],
                "matched_records": result["matched"],
                "match_percent": round(result["match_pct"], 2),
                "matched_columns": result["matched_columns"],
                "total_columns_compared": result["total_columns"],
                "column_match_ratio": round(result["matched_columns"] / result["total_columns"] * 100, 2)
                if result["total_columns"] else 0
            })

            all_sheets[name[:31]] = result["summary_df"].to_pandas()
            if INCLUDE_ROW_DIFFS and result["row_diff_df"].shape[0] > 0:
                all_sheets[f"{name[:28]}_diffs"] = result["row_diff_df"].to_pandas()
            if not result["financial_df"].empty:
                all_sheets[f"{name[:26]}_financials"] = result["financial_df"]
            if not result["column_stats_df"].empty:
                all_sheets[f"{name[:25]}_column_stats"] = result["column_stats_df"]
            if not result["false_mismatch_df"].empty:
                combined_false_mismatches = pd.concat([combined_false_mismatches, result["false_mismatch_df"]])

        except Exception as e:
            print(f"[{pair['name']}] ❌ Error: {e}")

    summary_df = pd.DataFrame(summary_records)
    all_sheets["summary"] = summary_df
    if not combined_false_mismatches.empty:
        all_sheets["false_mismatch_audit"] = combined_false_mismatches

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mismatch_summary_{model.lower()}_{timestamp}.xlsx" if TIMESTAMPED_REPORT else "mismatch_summary_report.xlsx"
    excel_path = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    apply_excel_formatting(excel_path)
    print(f"\n📊 Excel report saved: {excel_path}")

    # Export all sheets to organized CSVs
    csv_export_dir = os.path.join(OUTPUT_DIR, f"csv_exports/{model.lower()}_{timestamp}")
    os.makedirs(csv_export_dir, exist_ok=True)
    for sheet_name, df in all_sheets.items():
        filename = f"{sheet_name[:31].lower().replace(' ', '_')}.csv"
        csv_path = os.path.join(csv_export_dir, filename)
        df.to_csv(csv_path, index=False)

    print(f"📂 Exported all report tabs as CSVs to: {csv_export_dir}")
    print(f"✅ Completed in {datetime.now() - datetime.strptime(timestamp, '%Y%m%d_%H%M%S')}")

if __name__ == "__main__":
    main()
