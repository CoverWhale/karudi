import polars as pl
import clickhouse_connect
import pandas as pd
from datetime import datetime
import os
import argparse
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from functools import reduce
import operator

from dq_table_map import build_table_pairs

# ------------------------- ARGUMENTS ------------------------- #

def parse_args():
    parser = argparse.ArgumentParser(description="Run DQ checks for a given model and optional date.")
    parser.add_argument("--model", required=True, choices=["WIP", "Quotes", "Policy"], help="Data model to check")
    parser.add_argument("--date", help="Date folder to run against (format: YYYY-MM-DD)")
    return parser.parse_args()

# ------------------------- CONFIG ------------------------- #

INCLUDE_ROW_DIFFS = True
TIMESTAMPED_REPORT = True
OUTPUT_DIR = "dq_checks/mismatch_reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

client = clickhouse_connect.get_client(host="localhost")

# ------------------------- CORE LOGIC ------------------------- #

def fetch_table(table_name):
    result = client.query(f"SELECT * FROM {table_name}")
    return pl.DataFrame(result.result_rows, schema=result.column_names)

def compare_tables(name, df_a, df_b, join_keys, financial_fields):
    for a_key, b_key in join_keys.items():
        if a_key != b_key:
            df_b = df_b.rename({b_key: a_key})

    df_merged = df_a.join(df_b, on=list(join_keys.keys()), how="left", suffix="_b")

    non_key_b_cols = [c for c in df_b.columns if c not in join_keys.values()]
    if non_key_b_cols:
        match_indicator_col = f"{non_key_b_cols[0]}_b"
        matched = df_merged.filter(~pl.col(match_indicator_col).is_null()).shape[0]
    else:
        any_key = list(join_keys.keys())[0]
        matched = df_merged.filter(~pl.col(f"{any_key}_b").is_null()).shape[0]

    total_a = df_a.shape[0]
    match_pct = (matched / total_a) * 100 if total_a else 0
    print(f"[{name}] ✅ {matched}/{total_a} ({match_pct:.2f}%) records matched.")

    mismatch_summary = []
    common_cols = [c for c in df_a.columns if c in df_b.columns and c not in join_keys.keys()]
    for col in common_cols:
        col_a = pl.col(col)
        col_b = pl.col(f"{col}_b")
        mismatch_count = df_merged.filter((col_a != col_b) | col_a.is_null() | col_b.is_null()).shape[0]
        mismatch_summary.append((col, mismatch_count))
    summary_df = pl.DataFrame(mismatch_summary, schema=["column_name", "mismatches"]).sort("mismatches", descending=True)

    row_diff_df = pl.DataFrame()
    if INCLUDE_ROW_DIFFS and common_cols:
        mismatch_filter = reduce(operator.or_, [
            (pl.col(c) != pl.col(f"{c}_b")) | pl.col(c).is_null() | pl.col(f"{c}_b").is_null()
            for c in common_cols
        ])
        row_diff_df = df_merged.filter(mismatch_filter)

    financial_rows = []
    for field in financial_fields:
        if field in df_merged.columns and f"{field}_b" in df_merged.columns:
            a_sum = df_merged[field].fill_null(0).sum()
            b_sum = df_merged[f"{field}_b"].fill_null(0).sum()
            diff = a_sum - b_sum
            pct_diff = (diff / b_sum * 100) if b_sum != 0 else None
            financial_rows.append({
                "field": field,
                "table_a_sum": round(a_sum, 2),
                "table_b_sum": round(b_sum, 2),
                "difference": round(diff, 2),
                "percent_difference": round(pct_diff, 2) if pct_diff is not None else "N/A"
            })
    financial_df = pd.DataFrame(financial_rows)

    return {
        "summary_df": summary_df,
        "row_diff_df": row_diff_df,
        "financial_df": financial_df,
        "matched": matched,
        "total_a": total_a,
        "match_pct": match_pct
    }

def apply_excel_formatting(filepath):
    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith("_financials"):
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if abs(cell.value) > 5:
                            cell.fill = PatternFill(start_color="FF9999", fill_type="solid")
                        elif abs(cell.value) > 1:
                            cell.fill = PatternFill(start_color="FFF799", fill_type="solid")
    wb.save(filepath)

# ------------------------- MAIN ------------------------- #

def main():
    args = parse_args()
    model = args.model
    date_folder = args.date or datetime.now().strftime("%Y-%m-%d")
    print(f"🚀 Starting DQ checks for model: `{model}`, date folder: `{date_folder}`")

    table_pairs = build_table_pairs(model)

    all_sheets = {}
    summary_records = []

    for pair in table_pairs:
        try:
            name = pair["name"]
            print(f"\n🔍 Comparing: {name}")
            df_a = fetch_table(pair["table_a"])
            df_b = fetch_table(pair["table_b"])
            result = compare_tables(name, df_a, df_b, pair["join_keys"], pair.get("financial_fields", []))

            summary_records.append({
                "table_name": name,
                "rows_in_cw_table": df_a.shape[0],
                "rows_in_solartis_table": df_b.shape[0],
                "matched_records": result["matched"],
                "match_percent": round(result["match_pct"], 2)
            })

            all_sheets[name[:31]] = result["summary_df"].to_pandas()
            if INCLUDE_ROW_DIFFS and result["row_diff_df"].shape[0] > 0:
                all_sheets[f"{name[:28]}_diffs"] = result["row_diff_df"].to_pandas()
            if not result["financial_df"].empty:
                all_sheets[f"{name[:26]}_financials"] = result["financial_df"]

        except Exception as e:
            print(f"[{pair['name']}] ❌ Error: {e}")

    summary_df = pd.DataFrame(summary_records)
    all_sheets["summary"] = summary_df

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mismatch_summary_{model.lower()}_{timestamp}.xlsx" if TIMESTAMPED_REPORT else "mismatch_summary_report.xlsx"
    excel_path = os.path.join(OUTPUT_DIR, filename)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    apply_excel_formatting(excel_path)
    print(f"\n📊 Excel report saved: {excel_path}")
    print(f"✅ Completed in {datetime.now() - datetime.strptime(timestamp, '%Y%m%d_%H%M%S')}")

if __name__ == "__main__":
    main()
